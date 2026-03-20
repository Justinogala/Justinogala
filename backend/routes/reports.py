"""
Incident Reports (IR) and Serious Occurrence Reports (SOR) routes.
Provides CRUD operations with role-based access control,
PDF/Excel export, email notifications, and auto-escalation.
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone, timedelta
from typing import Optional, List
from pydantic import BaseModel
import uuid
import io
import asyncio
import resend

from config import db, logger, SENDER_EMAIL
from routes.chat import sse_manager

router = APIRouter(prefix="/reports", tags=["Incident Reports"])


# ============== IR/SOR Templates ==============

DEFAULT_IR_TEMPLATES = [
    {
        "id": "ir-tpl-workplace-injury",
        "name": "Workplace Injury",
        "category": "Injury",
        "description": "Report a physical injury sustained at the workplace",
        "icon": "hard-hat",
        "default_severity": "moderate",
        "is_default": True,
        "fields": [
            {"name": "injury_type", "label": "Type of Injury", "type": "select", "required": True, "options": ["Laceration", "Fracture", "Burn", "Sprain/Strain", "Concussion", "Chemical Exposure", "Other"]},
            {"name": "body_part", "label": "Body Part Affected", "type": "text", "required": True},
            {"name": "treatment_given", "label": "Treatment Given On-Site", "type": "textarea", "required": False},
            {"name": "medical_attention", "label": "Medical Attention Required", "type": "select", "required": True, "options": ["None", "First Aid", "Hospital Visit", "Ongoing Treatment"]},
            {"name": "equipment_involved", "label": "Equipment Involved", "type": "text", "required": False},
        ],
    },
    {
        "id": "ir-tpl-medication-error",
        "name": "Medication Error",
        "category": "Medical",
        "description": "Report a medication administration or dispensing error",
        "icon": "pill",
        "default_severity": "major",
        "is_default": True,
        "fields": [
            {"name": "medication_name", "label": "Medication Name", "type": "text", "required": True},
            {"name": "error_type", "label": "Error Type", "type": "select", "required": True, "options": ["Wrong Dose", "Wrong Medication", "Wrong Patient", "Wrong Route", "Wrong Time", "Omission", "Other"]},
            {"name": "prescribed_dose", "label": "Prescribed Dose", "type": "text", "required": True},
            {"name": "administered_dose", "label": "Dose Actually Given", "type": "text", "required": True},
            {"name": "adverse_effects", "label": "Adverse Effects Observed", "type": "textarea", "required": False},
            {"name": "physician_notified", "label": "Physician Notified", "type": "select", "required": True, "options": ["Yes", "No", "N/A"]},
        ],
    },
    {
        "id": "ir-tpl-property-damage",
        "name": "Property Damage",
        "category": "Property",
        "description": "Report damage to facility property or equipment",
        "icon": "building",
        "default_severity": "moderate",
        "is_default": True,
        "fields": [
            {"name": "property_type", "label": "Property Type", "type": "select", "required": True, "options": ["Building/Structure", "Equipment", "Vehicle", "Furniture", "IT/Electronics", "Other"]},
            {"name": "damage_description", "label": "Damage Description", "type": "textarea", "required": True},
            {"name": "estimated_cost", "label": "Estimated Repair Cost", "type": "number", "required": False},
            {"name": "insurance_claim", "label": "Insurance Claim Required", "type": "select", "required": True, "options": ["Yes", "No", "Undetermined"]},
        ],
    },
    {
        "id": "ir-tpl-behavioural",
        "name": "Behavioural Incident",
        "category": "Behavioural",
        "description": "Report aggressive, disruptive, or threatening behaviour",
        "icon": "alert-triangle",
        "default_severity": "moderate",
        "is_default": True,
        "fields": [
            {"name": "behaviour_type", "label": "Behaviour Type", "type": "select", "required": True, "options": ["Verbal Aggression", "Physical Aggression", "Self-Harm", "Elopement", "Property Destruction", "Disruptive Behaviour", "Other"]},
            {"name": "trigger", "label": "Known Trigger / Antecedent", "type": "textarea", "required": False},
            {"name": "intervention_used", "label": "Intervention Used", "type": "textarea", "required": True},
            {"name": "restraint_used", "label": "Restraint Used", "type": "select", "required": True, "options": ["None", "Physical", "Chemical", "Mechanical", "Seclusion"]},
            {"name": "outcome", "label": "Outcome", "type": "textarea", "required": True},
        ],
    },
    {
        "id": "ir-tpl-safeguarding",
        "name": "Safeguarding Concern",
        "category": "Safeguarding",
        "description": "Report a safeguarding or child/adult protection concern",
        "icon": "shield",
        "default_severity": "major",
        "is_default": True,
        "fields": [
            {"name": "concern_type", "label": "Type of Concern", "type": "select", "required": True, "options": ["Physical Abuse", "Emotional Abuse", "Sexual Abuse", "Neglect", "Financial Exploitation", "Self-Neglect", "Other"]},
            {"name": "vulnerable_person", "label": "Vulnerable Person Category", "type": "select", "required": True, "options": ["Child", "Elderly Adult", "Adult with Disabilities", "Other"]},
            {"name": "disclosure", "label": "Was There a Disclosure?", "type": "select", "required": True, "options": ["Yes — verbal", "Yes — written", "No, observed signs", "Third party report"]},
            {"name": "authority_notified", "label": "Authority Notified", "type": "select", "required": True, "options": ["Not yet", "CAS/CPS", "Police", "Adult Protection", "Other"]},
            {"name": "safety_plan", "label": "Immediate Safety Plan", "type": "textarea", "required": True},
        ],
    },
    {
        "id": "ir-tpl-near-miss",
        "name": "Near Miss",
        "category": "Safety",
        "description": "Report an event that could have caused harm but did not",
        "icon": "zap",
        "default_severity": "minor",
        "is_default": True,
        "fields": [
            {"name": "hazard_type", "label": "Hazard Type", "type": "select", "required": True, "options": ["Slip/Trip/Fall", "Equipment Malfunction", "Chemical Spill", "Electrical", "Fire Risk", "Ergonomic", "Other"]},
            {"name": "potential_outcome", "label": "Potential Outcome If Not Caught", "type": "textarea", "required": True},
            {"name": "contributing_factors", "label": "Contributing Factors", "type": "textarea", "required": True},
            {"name": "recommended_action", "label": "Recommended Preventive Action", "type": "textarea", "required": True},
        ],
    },
    {
        "id": "ir-tpl-serious-occurrence",
        "name": "Serious Occurrence",
        "category": "SOR",
        "description": "Report a serious occurrence requiring mandatory regulatory notification",
        "icon": "alert-octagon",
        "default_severity": "serious_occurrence",
        "is_default": True,
        "fields": [
            {"name": "sor_category", "label": "SOR Category", "type": "select", "required": True, "options": ["Death", "Serious Injury", "Abuse/Neglect", "Missing Person", "Disaster", "Outbreak", "Unauthorized Absence", "Complaint with Regulatory Body", "Other"]},
            {"name": "regulatory_body", "label": "Regulatory Body to Notify", "type": "text", "required": True},
            {"name": "notification_deadline", "label": "Notification Deadline", "type": "text", "required": True},
            {"name": "notification_sent", "label": "Notification Sent", "type": "select", "required": True, "options": ["Yes", "No", "In progress"]},
            {"name": "family_notified", "label": "Family/Guardian Notified", "type": "select", "required": True, "options": ["Yes", "No", "N/A"]},
            {"name": "immediate_actions", "label": "Immediate Actions Taken", "type": "textarea", "required": True},
        ],
    },
]

IR_TEMPLATE_CATEGORIES = ["Injury", "Medical", "Property", "Behavioural", "Safeguarding", "Safety", "SOR", "Custom"]


class CreateIRTemplate(BaseModel):
    name: str
    category: str
    description: str = ""
    icon: str = "file-text"
    default_severity: str = "moderate"
    fields: List[dict] = []


@router.get("/templates")
async def get_ir_templates(category: Optional[str] = None):
    """Get all IR/SOR templates (default + custom)."""
    templates = list(DEFAULT_IR_TEMPLATES)
    custom = await db.ir_sor_templates.find({}, {"_id": 0}).to_list(500)
    templates.extend(custom)
    if category:
        templates = [t for t in templates if t.get("category") == category]
    return {"templates": templates, "categories": IR_TEMPLATE_CATEGORIES}


@router.get("/templates/{template_id}")
async def get_ir_template(template_id: str):
    """Get a specific IR/SOR template by ID."""
    for t in DEFAULT_IR_TEMPLATES:
        if t["id"] == template_id:
            return t
    custom = await db.ir_sor_templates.find_one({"id": template_id}, {"_id": 0})
    if custom:
        return custom
    raise HTTPException(status_code=404, detail="Template not found")


@router.post("/templates")
async def create_ir_template(req: CreateIRTemplate):
    """Create a custom IR/SOR template (superadmin only)."""
    template = {
        "id": f"ir-tpl-custom-{uuid.uuid4().hex[:8]}",
        "name": req.name,
        "category": req.category,
        "description": req.description,
        "icon": req.icon,
        "default_severity": req.default_severity,
        "is_default": False,
        "fields": req.fields,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.ir_sor_templates.insert_one(template)
    return {"success": True, "template": {k: v for k, v in template.items() if k != "_id"}}


@router.put("/templates/{template_id}")
async def update_ir_template(template_id: str, req: CreateIRTemplate):
    """Update a custom IR/SOR template."""
    # Block editing default templates
    for t in DEFAULT_IR_TEMPLATES:
        if t["id"] == template_id:
            raise HTTPException(status_code=400, detail="Cannot edit default templates")

    result = await db.ir_sor_templates.update_one(
        {"id": template_id},
        {"$set": {
            "name": req.name,
            "category": req.category,
            "description": req.description,
            "icon": req.icon,
            "default_severity": req.default_severity,
            "fields": req.fields,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")

    updated = await db.ir_sor_templates.find_one({"id": template_id}, {"_id": 0})
    return {"success": True, "template": updated}


@router.delete("/templates/{template_id}")
async def delete_ir_template(template_id: str):
    """Delete a custom IR/SOR template."""
    for t in DEFAULT_IR_TEMPLATES:
        if t["id"] == template_id:
            raise HTTPException(status_code=400, detail="Cannot delete default templates")

    result = await db.ir_sor_templates.delete_one({"id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"success": True}


# ============== Models ==============

class PersonInvolved(BaseModel):
    full_name: str
    role: str  # staff, client, visitor
    contact_info: Optional[str] = None

class ReportCreate(BaseModel):
    workspace_id: str
    submitted_by: str  # user ID
    template_id: Optional[str] = None
    template_name: Optional[str] = None
    custom_fields: Optional[dict] = None  # template-specific fields
    # Section A - Incident Details
    incident_date: str
    incident_time: str
    location: str
    department: Optional[str] = None
    incident_type: str  # injury, medication_error, property_damage, behavioural, safeguarding, near_miss, other
    incident_type_other: Optional[str] = None
    # Section B - Persons Involved
    persons_involved: List[PersonInvolved] = []
    witnesses: Optional[str] = None
    # Section C - Description
    description: str
    immediate_action: Optional[str] = None
    was_911_called: bool = False
    # Section D - Severity
    severity: str  # minor, moderate, major, critical, serious_occurrence
    # Section F - Follow-up (optional on creation)
    assigned_investigator: Optional[str] = None

class ReportUpdate(BaseModel):
    # Section A
    incident_date: Optional[str] = None
    incident_time: Optional[str] = None
    location: Optional[str] = None
    department: Optional[str] = None
    incident_type: Optional[str] = None
    incident_type_other: Optional[str] = None
    # Section B
    persons_involved: Optional[List[PersonInvolved]] = None
    witnesses: Optional[str] = None
    # Section C
    description: Optional[str] = None
    immediate_action: Optional[str] = None
    was_911_called: Optional[bool] = None
    # Section D
    severity: Optional[str] = None
    # Section F - Investigation (manager/admin only)
    assigned_investigator: Optional[str] = None
    root_cause: Optional[str] = None
    corrective_action: Optional[str] = None
    follow_up_due_date: Optional[str] = None
    status: Optional[str] = None  # open, under_review, closed
    investigation_notes: Optional[str] = None


# ============== Endpoints ==============

@router.post("")
async def create_report(report: ReportCreate):
    """Create a new incident report. Any authenticated user can submit."""
    try:
        report_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc).isoformat()
        
        # Get submitter info
        submitter = await db.users.find_one({"id": report.submitted_by}, {"_id": 0, "name": 1, "email": 1, "role": 1})
        
        # Determine report type
        report_type = "SOR" if report.severity == "serious_occurrence" else "IR"
        
        doc = {
            "id": report_id,
            "report_number": f"{report_type}-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{report_id[:6].upper()}",
            "report_type": report_type,
            "workspace_id": report.workspace_id,
            "submitted_by": report.submitted_by,
            "submitted_by_name": submitter.get("name") if submitter else None,
            "submitted_by_email": submitter.get("email") if submitter else None,
            # Section A
            "incident_date": report.incident_date,
            "incident_time": report.incident_time,
            "location": report.location,
            "department": report.department,
            "incident_type": report.incident_type,
            "incident_type_other": report.incident_type_other,
            # Section B
            "persons_involved": [p.dict() for p in report.persons_involved],
            "witnesses": report.witnesses,
            # Section C
            "description": report.description,
            "immediate_action": report.immediate_action,
            "was_911_called": report.was_911_called,
            # Section D
            "severity": report.severity,
            # Section E
            "attachments": [],
            # Section F
            "assigned_investigator": report.assigned_investigator,
            "root_cause": None,
            "corrective_action": None,
            "follow_up_due_date": None,
            "investigation_notes": None,
            "status": "open",
            # Metadata
            "template_id": report.template_id,
            "template_name": report.template_name,
            "custom_fields": report.custom_fields,
            "created_at": now,
            "updated_at": now,
            "audit_log": [{
                "action": "created",
                "by": report.submitted_by,
                "by_name": submitter.get("name") if submitter else None,
                "at": now,
            }]
        }
        
        await db.incident_reports.insert_one(doc)
        doc.pop("_id", None)
        
        # Send email notification for critical/SOR incidents (fire and forget)
        if report.severity in ("critical", "serious_occurrence"):
            asyncio.create_task(_notify_critical_incident(doc))
            asyncio.create_task(_broadcast_critical_incident_sse(doc))
        
        return {"success": True, "report": doc}
    except Exception as e:
        logger.error(f"Error creating report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("")
async def list_reports(
    workspace_id: Optional[str] = None,
    user_id: Optional[str] = None,
    user_role: Optional[str] = None,
    report_type: Optional[str] = None,  # IR, SOR
    severity: Optional[str] = None,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
):
    """List reports with role-based filtering."""
    try:
        query = {}
        
        if workspace_id:
            query["workspace_id"] = workspace_id
        
        # Role-based access: staff can only see their own
        if user_role and user_role not in ["Admin", "Manager"] and user_id:
            query["submitted_by"] = user_id
        
        if report_type:
            query["report_type"] = report_type
        if severity:
            query["severity"] = severity
        if status:
            query["status"] = status
        
        total = await db.incident_reports.count_documents(query)
        skip = (page - 1) * limit
        
        reports = await db.incident_reports.find(
            query, {"_id": 0}
        ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
        
        return {
            "success": True,
            "reports": reports,
            "total": total,
            "page": page,
            "pages": (total + limit - 1) // limit
        }
    except Exception as e:
        logger.error(f"Error listing reports: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stats")
async def get_report_stats(workspace_id: Optional[str] = None):
    """Dashboard statistics for reports."""
    try:
        query = {}
        if workspace_id:
            query["workspace_id"] = workspace_id
        
        total = await db.incident_reports.count_documents(query)
        open_count = await db.incident_reports.count_documents({**query, "status": "open"})
        review_count = await db.incident_reports.count_documents({**query, "status": "under_review"})
        closed_count = await db.incident_reports.count_documents({**query, "status": "closed"})
        
        ir_count = await db.incident_reports.count_documents({**query, "report_type": "IR"})
        sor_count = await db.incident_reports.count_documents({**query, "report_type": "SOR"})
        
        critical_count = await db.incident_reports.count_documents({**query, "severity": {"$in": ["critical", "serious_occurrence"]}})
        
        # By type
        by_type = {}
        for itype in ["injury", "medication_error", "property_damage", "behavioural", "safeguarding", "near_miss", "other"]:
            c = await db.incident_reports.count_documents({**query, "incident_type": itype})
            if c > 0:
                by_type[itype] = c
        
        # By severity
        by_severity = {}
        for sev in ["minor", "moderate", "major", "critical", "serious_occurrence"]:
            c = await db.incident_reports.count_documents({**query, "severity": sev})
            if c > 0:
                by_severity[sev] = c
        
        return {
            "success": True,
            "stats": {
                "total": total,
                "open": open_count,
                "under_review": review_count,
                "closed": closed_count,
                "ir_count": ir_count,
                "sor_count": sor_count,
                "critical": critical_count,
                "by_type": by_type,
                "by_severity": by_severity,
            }
        }
    except Exception as e:
        logger.error(f"Error getting report stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/analytics")
async def get_report_analytics(workspace_id: Optional[str] = None):
    """Advanced analytics: severity trends, type breakdown, response time metrics."""
    try:
        match_stage = {}
        if workspace_id:
            match_stage["workspace_id"] = workspace_id

        # 1) Severity trend by month (last 12 months)
        severity_trend_pipeline = [
            {"$match": match_stage} if match_stage else {"$match": {}},
            {"$addFields": {
                "month": {"$substr": ["$created_at", 0, 7]},
            }},
            {"$group": {
                "_id": {"month": "$month", "severity": "$severity"},
                "count": {"$sum": 1},
            }},
            {"$sort": {"_id.month": 1}},
        ]
        severity_trend_raw = await db.incident_reports.aggregate(severity_trend_pipeline).to_list(500)

        months_set = sorted({r["_id"]["month"] for r in severity_trend_raw})
        months_set = months_set[-12:]  # last 12 months
        severity_keys = ["minor", "moderate", "major", "critical", "serious_occurrence"]
        severity_trend = []
        for m in months_set:
            entry = {"month": m}
            for sev in severity_keys:
                entry[sev] = 0
            for r in severity_trend_raw:
                if r["_id"]["month"] == m and r["_id"]["severity"] in severity_keys:
                    entry[r["_id"]["severity"]] = r["count"]
            severity_trend.append(entry)

        # 2) Incident type breakdown
        type_pipeline = [
            {"$match": match_stage} if match_stage else {"$match": {}},
            {"$group": {"_id": "$incident_type", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
        ]
        type_breakdown = [
            {"type": r["_id"], "count": r["count"]}
            for r in await db.incident_reports.aggregate(type_pipeline).to_list(20)
        ]

        # 3) Response time: avg hours from created_at to first status change (audit_log)
        all_reports = await db.incident_reports.find(
            match_stage if match_stage else {},
            {"_id": 0, "severity": 1, "status": 1, "created_at": 1, "audit_log": 1}
        ).to_list(5000)

        from datetime import datetime as dt
        response_times = {sev: [] for sev in severity_keys}
        for r in all_reports:
            sev = r.get("severity", "minor")
            if sev not in response_times:
                continue
            created_str = r.get("created_at", "")
            audit = r.get("audit_log", [])
            review_entry = next(
                (e for e in audit if e.get("action") == "updated" and "status" in (e.get("fields_changed") or [])),
                None
            )
            if review_entry and created_str:
                try:
                    created_dt = dt.fromisoformat(created_str.replace("Z", "+00:00"))
                    reviewed_dt = dt.fromisoformat(review_entry["at"].replace("Z", "+00:00"))
                    hours = (reviewed_dt - created_dt).total_seconds() / 3600
                    if hours >= 0:
                        response_times[sev].append(round(hours, 1))
                except Exception:
                    pass

        avg_response = []
        for sev in severity_keys:
            vals = response_times[sev]
            avg_response.append({
                "severity": sev,
                "avg_hours": round(sum(vals) / len(vals), 1) if vals else None,
                "count": len(vals),
            })

        # 4) Monthly summary (current month vs previous month)
        now = datetime.now(timezone.utc)
        current_month = now.strftime("%Y-%m")
        prev_month_dt = (now.replace(day=1) - timedelta(days=1))
        prev_month = prev_month_dt.strftime("%Y-%m")

        current_count = 0
        prev_count = 0
        current_closed = 0
        for r in all_reports:
            m = (r.get("created_at") or "")[:7]
            if m == current_month:
                current_count += 1
                if r.get("status") == "closed":
                    current_closed += 1
            elif m == prev_month:
                prev_count += 1

        escalated_count = await db.incident_reports.count_documents({
            **(match_stage if match_stage else {}),
            "escalated": True,
            "created_at": {"$gte": f"{current_month}-01"},
        })

        monthly_summary = {
            "current_month": current_month,
            "current_count": current_count,
            "prev_count": prev_count,
            "change_pct": round(((current_count - prev_count) / prev_count * 100) if prev_count else 0, 1),
            "closure_rate": round((current_closed / current_count * 100) if current_count else 0, 1),
            "escalated": escalated_count,
        }

        return {
            "success": True,
            "analytics": {
                "severity_trend": severity_trend,
                "type_breakdown": type_breakdown,
                "response_times": avg_response,
                "monthly_summary": monthly_summary,
            }
        }
    except Exception as e:
        logger.error(f"Error getting analytics: {e}")
        raise HTTPException(status_code=500, detail=str(e))




# ============== Excel Export ==============

@router.get("/export/excel")
async def export_reports_excel(
    workspace_id: Optional[str] = None,
    report_type: Optional[str] = None,
    severity: Optional[str] = None,
    status: Optional[str] = None,
):
    """Export filtered reports as an Excel spreadsheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        query = {}
        if workspace_id:
            query["workspace_id"] = workspace_id
        if report_type:
            query["report_type"] = report_type
        if severity:
            query["severity"] = severity
        if status:
            query["status"] = status

        reports = await db.incident_reports.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)

        wb = Workbook()
        ws = wb.active
        ws.title = "Incident Reports"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )

        headers = [
            "Report #", "Type", "Date", "Time", "Location", "Department",
            "Incident Type", "Severity", "Status", "Submitted By",
            "Description", "Immediate Action", "911 Called",
            "Persons Involved", "Witnesses",
            "Investigator", "Root Cause", "Corrective Action", "Due Date",
            "Created At"
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        severity_fills = {
            "minor": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
            "moderate": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
            "major": PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
            "critical": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            "serious_occurrence": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
        }

        for row_num, r in enumerate(reports, 2):
            persons_str = "; ".join(f"{p.get('full_name','')} ({p.get('role','')})" for p in r.get("persons_involved", []))
            row_data = [
                r.get("report_number", ""),
                r.get("report_type", ""),
                r.get("incident_date", ""),
                r.get("incident_time", ""),
                r.get("location", ""),
                r.get("department", ""),
                r.get("incident_type", "").replace("_", " ").title(),
                r.get("severity", "").replace("_", " ").title(),
                r.get("status", "").replace("_", " ").title(),
                r.get("submitted_by_name", ""),
                r.get("description", ""),
                r.get("immediate_action", ""),
                "Yes" if r.get("was_911_called") else "No",
                persons_str,
                r.get("witnesses", ""),
                r.get("assigned_investigator", ""),
                r.get("root_cause", ""),
                r.get("corrective_action", ""),
                r.get("follow_up_due_date", ""),
                r.get("created_at", ""),
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value or "")
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            sev = r.get("severity", "")
            if sev in severity_fills:
                ws.cell(row=row_num, column=8).fill = severity_fills[sev]

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 40))
            ws.column_dimensions[col_letter].width = max_length + 3

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"incident_reports_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        logger.error(f"Error exporting Excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============== Bulk PDF Export ==============

@router.get("/export/pdf")
async def export_reports_bulk_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    severity: Optional[str] = None,
    status: Optional[str] = None,
):
    """Export all/filtered incident reports as a single multi-page PDF."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        query = {}
        if start_date:
            query.setdefault("created_at", {})["$gte"] = start_date
        if end_date:
            query.setdefault("created_at", {})["$lte"] = end_date
        if severity:
            query["severity"] = severity
        if status:
            query["status"] = status

        reports = await db.incident_reports.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle('CoverTitle', parent=styles['Title'], fontSize=22, textColor=colors.HexColor('#1e1b4b'), spaceAfter=6)
        subtitle_style = ParagraphStyle('CoverSub', parent=styles['Normal'], fontSize=11, textColor=colors.grey, spaceAfter=20)
        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#4338ca'), spaceBefore=10, spaceAfter=4)
        body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=9, leading=12, spaceAfter=4)
        header_style = ParagraphStyle('ReportHeader', parent=styles['Heading3'], fontSize=13, textColor=colors.HexColor('#1e1b4b'), spaceBefore=4, spaceAfter=2)

        # Cover page
        story.append(Spacer(1, 40))
        story.append(Paragraph("Incident Reports — Summary", title_style))
        date_range = ""
        if start_date and end_date:
            date_range = f"{start_date} to {end_date}"
        elif start_date:
            date_range = f"From {start_date}"
        elif end_date:
            date_range = f"Up to {end_date}"
        else:
            date_range = "All records"
        story.append(Paragraph(f"{len(reports)} reports &bull; {date_range}", subtitle_style))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb')))
        story.append(Spacer(1, 10))

        # Summary table
        sev_counts = {}
        for r in reports:
            s = r.get("severity", "unknown")
            sev_counts[s] = sev_counts.get(s, 0) + 1

        summary_data = [["Severity", "Count"]]
        for s in ["minor", "moderate", "major", "critical", "serious_occurrence"]:
            if s in sev_counts:
                summary_data.append([s.replace("_", " ").title(), str(sev_counts[s])])
        summary_data.append(["Total", str(len(reports))])

        t = Table(summary_data, colWidths=[200, 80])
        t.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4338ca')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f3f4f6')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        story.append(t)

        # Individual reports
        for i, r in enumerate(reports):
            if i > 0:
                story.append(PageBreak())

            report_type = "SOR" if r.get("report_type") == "SOR" else "IR"
            story.append(header_style and Paragraph(f"{r.get('report_number', '')} — {report_type}", header_style))
            story.append(Paragraph(f"Status: {(r.get('status') or 'open').replace('_', ' ').title()} &bull; Severity: {r.get('severity', '').replace('_', ' ').title()}", body_style))
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#e5e7eb')))

            details = [
                ["Date", r.get("incident_date", ""), "Time", r.get("incident_time", "")],
                ["Location", r.get("location", ""), "Dept", r.get("department", "") or "—"],
                ["Type", r.get("incident_type", "").replace("_", " ").title(), "Submitted", r.get("submitted_by_name", "")],
            ]
            dt = Table(details, colWidths=[50, 150, 50, 150])
            dt.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
                ('TEXTCOLOR', (2, 0), (2, -1), colors.grey),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
            ]))
            story.append(dt)

            if r.get("description"):
                story.append(Paragraph("Description", section_style))
                story.append(Paragraph(r["description"][:500], body_style))

            if r.get("immediate_action"):
                story.append(Paragraph("Immediate Action", section_style))
                story.append(Paragraph(r["immediate_action"][:300], body_style))

            persons = r.get("persons_involved", [])
            if persons:
                story.append(Paragraph("Persons Involved", section_style))
                for p in persons:
                    story.append(Paragraph(f"&bull; {p.get('full_name', '')} ({p.get('role', '')})", body_style))

        # Footer
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#d1d5db')))
        story.append(Paragraph(f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} — Munal AI Incident Reporting", ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.lightgrey, spaceBefore=4)))

        doc.build(story)
        buf.seek(0)

        filename = f"incident_reports_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.pdf"
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        logger.error(f"Error exporting bulk PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{report_id}")
async def get_report(report_id: str):
    """Get a single report by ID."""
    try:
        report = await db.incident_reports.find_one({"id": report_id}, {"_id": 0})
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")
        return {"success": True, "report": report}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{report_id}")
async def update_report(report_id: str, update: ReportUpdate, editor_id: str = None):
    """Update a report. Manager/Admin can update status and investigation fields."""
    try:
        existing = await db.incident_reports.find_one({"id": report_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Report not found")
        
        update_data = {k: v for k, v in update.dict().items() if v is not None}
        if not update_data:
            raise HTTPException(status_code=400, detail="No fields to update")
        
        # Handle persons_involved serialization
        if "persons_involved" in update_data:
            update_data["persons_involved"] = [p.dict() if hasattr(p, 'dict') else p for p in update_data["persons_involved"]]
        
        now = datetime.now(timezone.utc).isoformat()
        update_data["updated_at"] = now
        
        # Auto-detect report type change
        if "severity" in update_data:
            update_data["report_type"] = "SOR" if update_data["severity"] == "serious_occurrence" else "IR"
        
        # Add audit log entry
        editor = None
        if editor_id:
            editor = await db.users.find_one({"id": editor_id}, {"_id": 0, "name": 1})
        
        audit_entry = {
            "action": "updated",
            "by": editor_id or "unknown",
            "by_name": editor.get("name") if editor else None,
            "at": now,
            "fields_changed": list(update_data.keys())
        }
        
        await db.incident_reports.update_one(
            {"id": report_id},
            {
                "$set": update_data,
                "$push": {"audit_log": audit_entry}
            }
        )
        
        updated = await db.incident_reports.find_one({"id": report_id}, {"_id": 0})
        return {"success": True, "report": updated}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{report_id}/attachments")
async def upload_report_attachment(
    report_id: str,
    file: UploadFile = File(...),
    user_id: str = Form(...)
):
    """Upload an attachment to a report."""
    try:
        report = await db.incident_reports.find_one({"id": report_id})
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")
        
        contents = await file.read()
        if len(contents) > 20 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File too large. Maximum 20MB.")
        
        attachment_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc).isoformat()
        
        # Store file in GridFS
        from motor.motor_asyncio import AsyncIOMotorGridFSBucket
        fs = AsyncIOMotorGridFSBucket(db, bucket_name="report_attachments")
        grid_id = await fs.upload_from_stream(
            file.filename,
            contents,
            metadata={
                "report_id": report_id,
                "attachment_id": attachment_id,
                "user_id": user_id,
                "content_type": file.content_type,
                "size": len(contents),
            }
        )
        
        attachment = {
            "id": attachment_id,
            "grid_id": str(grid_id),
            "filename": file.filename,
            "content_type": file.content_type,
            "size": len(contents),
            "uploaded_by": user_id,
            "uploaded_at": now,
        }
        
        await db.incident_reports.update_one(
            {"id": report_id},
            {"$push": {"attachments": attachment}}
        )
        
        return {"success": True, "attachment": attachment}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading report attachment: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{report_id}/attachments/{attachment_id}")
async def download_report_attachment(report_id: str, attachment_id: str):
    """Download a report attachment."""
    try:
        from bson import ObjectId
        from fastapi.responses import StreamingResponse
        from motor.motor_asyncio import AsyncIOMotorGridFSBucket
        
        report = await db.incident_reports.find_one({"id": report_id}, {"_id": 0, "attachments": 1})
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")
        
        attachment = next((a for a in report.get("attachments", []) if a["id"] == attachment_id), None)
        if not attachment:
            raise HTTPException(status_code=404, detail="Attachment not found")
        
        fs = AsyncIOMotorGridFSBucket(db, bucket_name="report_attachments")
        grid_out = await fs.open_download_stream(ObjectId(attachment["grid_id"]))
        
        async def file_stream():
            while True:
                chunk = await grid_out.read(8192)
                if not chunk:
                    break
                yield chunk
        
        return StreamingResponse(
            file_stream(),
            media_type=attachment.get("content_type", "application/octet-stream"),
            headers={
                "Content-Disposition": f'attachment; filename="{attachment["filename"]}"',
                "Content-Length": str(attachment.get("size", 0))
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading attachment: {e}")
        raise HTTPException(status_code=500, detail=str(e))



# ============== PDF Export ==============

@router.get("/{report_id}/export/pdf")
async def export_report_pdf(report_id: str):
    """Export a single report as a PDF document."""
    try:
        report = await db.incident_reports.find_one({"id": report_id}, {"_id": 0})
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")

        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm, leftMargin=20*mm, rightMargin=20*mm)
        styles = getSampleStyleSheet()
        story = []

        # Custom styles
        title_style = ParagraphStyle('ReportTitle', parent=styles['Title'], fontSize=18, spaceAfter=4, textColor=colors.HexColor('#1e1b4b'))
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.grey, spaceAfter=12)
        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor('#4338ca'), spaceBefore=14, spaceAfter=6)
        body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=10, leading=14, spaceAfter=8)

        # Header
        report_type_label = "Serious Occurrence Report" if report.get("report_type") == "SOR" else "Incident Report"
        story.append(Paragraph(f"{report_type_label}", title_style))
        story.append(Paragraph(f"{report.get('report_number', 'N/A')} &bull; Status: {(report.get('status') or 'open').replace('_', ' ').title()}", subtitle_style))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb')))
        story.append(Spacer(1, 6))

        # Section A
        story.append(Paragraph("Section A — Incident Details", section_style))
        severity_label = report.get("severity", "").replace("_", " ").title()
        incident_type_label = report.get("incident_type", "").replace("_", " ").title()
        details_data = [
            ["Date", report.get("incident_date", ""), "Time", report.get("incident_time", "")],
            ["Location", report.get("location", ""), "Department", report.get("department", "") or "—"],
            ["Type", incident_type_label, "Severity", severity_label],
        ]
        t = Table(details_data, colWidths=[60, 170, 60, 170])
        t.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (2, 0), (2, -1), colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(t)

        # Section B
        story.append(Paragraph("Section B — Person(s) Involved", section_style))
        for i, person in enumerate(report.get("persons_involved", [])):
            story.append(Paragraph(f"<b>{i+1}.</b> {person.get('full_name', '')} ({person.get('role', '')}) {person.get('contact_info', '') or ''}", body_style))
        if report.get("witnesses"):
            story.append(Paragraph(f"<b>Witnesses:</b> {report['witnesses']}", body_style))

        # Section C
        story.append(Paragraph("Section C — Description", section_style))
        story.append(Paragraph(report.get("description", ""), body_style))
        if report.get("immediate_action"):
            story.append(Paragraph(f"<b>Immediate Action:</b> {report['immediate_action']}", body_style))
        story.append(Paragraph(f"<b>911 Called:</b> {'Yes' if report.get('was_911_called') else 'No'}", body_style))

        # Section F
        if any(report.get(k) for k in ("assigned_investigator", "root_cause", "corrective_action", "investigation_notes")):
            story.append(Paragraph("Section F — Investigation & Follow-Up", section_style))
            for label, key in [("Investigator", "assigned_investigator"), ("Root Cause", "root_cause"), ("Corrective Action", "corrective_action"), ("Due Date", "follow_up_due_date"), ("Notes", "investigation_notes")]:
                val = report.get(key)
                if val:
                    story.append(Paragraph(f"<b>{label}:</b> {val}", body_style))

        # Audit trail
        if report.get("audit_log"):
            story.append(Paragraph("Audit Trail", section_style))
            for entry in report["audit_log"]:
                ts = entry.get("at", "")
                story.append(Paragraph(f"&bull; <b>{entry.get('action', '').title()}</b> by {entry.get('by_name') or entry.get('by', '')} — {ts}", ParagraphStyle('AuditEntry', parent=styles['Normal'], fontSize=8, textColor=colors.grey, spaceAfter=2)))

        # Footer
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#d1d5db')))
        story.append(Paragraph(f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} — Munal AI Incident Reporting", ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.lightgrey, spaceBefore=4)))

        doc.build(story)
        buf.seek(0)

        filename = f"{report.get('report_number', 'report')}.pdf"
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============== SSE Real-Time Notification ==============

async def _broadcast_critical_incident_sse(report: dict):
    """Push a real-time notification to all connected admins/managers via SSE."""
    try:
        admins = await db.users.find(
            {"role": {"$in": ["Admin", "Manager"]}},
            {"_id": 0, "id": 1}
        ).to_list(100)

        severity_label = report.get("severity", "").replace("_", " ").title()
        payload = {
            "report_id": report.get("id"),
            "report_number": report.get("report_number"),
            "severity": report.get("severity"),
            "severity_label": severity_label,
            "report_type": report.get("report_type"),
            "location": report.get("location"),
            "submitted_by_name": report.get("submitted_by_name"),
            "incident_type": report.get("incident_type", "").replace("_", " ").title(),
            "timestamp": report.get("created_at"),
        }

        sent = 0
        for admin in admins:
            uid = admin.get("id")
            if uid and sse_manager.is_user_online(uid):
                await sse_manager.send_to_user(uid, "critical_incident", payload)
                sent += 1

        logger.info(f"SSE critical_incident broadcast sent to {sent} online admin/managers for {report.get('report_number')}")
    except Exception as e:
        logger.error(f"SSE critical_incident broadcast failed: {e}")


# ============== Email Notification Helpers ==============

async def _notify_critical_incident(report: dict):
    """Send email notification to admins/managers for critical or SOR incidents."""
    try:
        admins = await db.users.find(
            {"role": {"$in": ["Admin", "Manager"]}},
            {"_id": 0, "email": 1, "name": 1}
        ).to_list(100)

        if not admins:
            logger.info("No admins/managers to notify for critical incident")
            return

        recipient_emails = [a["email"] for a in admins if a.get("email")]
        if not recipient_emails:
            return

        severity_label = report.get("severity", "").replace("_", " ").title()
        report_type = "Serious Occurrence Report (SOR)" if report.get("report_type") == "SOR" else "Incident Report"

        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px;">
            <div style="text-align:center;padding:16px 0;">
                <h1 style="color:#7c3aed;margin:0;">Munal AI</h1>
            </div>
            <div style="background:#fef2f2;border-left:4px solid #dc2626;border-radius:8px;padding:24px;margin:16px 0;">
                <h2 style="color:#991b1b;margin-top:0;">Critical Incident Alert</h2>
                <p style="color:#4b5563;">A <strong>{severity_label}</strong> {report_type} has been submitted and requires your immediate attention.</p>
                <table style="width:100%;border-collapse:collapse;margin:16px 0;">
                    <tr><td style="padding:6px 0;color:#6b7280;width:130px;">Report #</td><td style="padding:6px 0;font-weight:bold;">{report.get('report_number', 'N/A')}</td></tr>
                    <tr><td style="padding:6px 0;color:#6b7280;">Date/Time</td><td style="padding:6px 0;">{report.get('incident_date', '')} at {report.get('incident_time', '')}</td></tr>
                    <tr><td style="padding:6px 0;color:#6b7280;">Location</td><td style="padding:6px 0;">{report.get('location', '')}</td></tr>
                    <tr><td style="padding:6px 0;color:#6b7280;">Submitted By</td><td style="padding:6px 0;">{report.get('submitted_by_name', 'Unknown')}</td></tr>
                    <tr><td style="padding:6px 0;color:#6b7280;">Type</td><td style="padding:6px 0;">{report.get('incident_type', '').replace('_', ' ').title()}</td></tr>
                </table>
                <p style="color:#4b5563;"><strong>Description:</strong><br/>{report.get('description', '')[:300]}{'...' if len(report.get('description', '')) > 300 else ''}</p>
            </div>
            <p style="color:#6b7280;font-size:13px;">Please review this report at your earliest convenience.</p>
            <div style="text-align:center;padding:16px 0;border-top:1px solid #e5e7eb;">
                <p style="color:#9ca3af;font-size:11px;">&copy; 2026 Munal AI. All rights reserved.</p>
            </div>
        </div>
        """

        params = {
            "from": SENDER_EMAIL,
            "to": recipient_emails,
            "subject": f"[CRITICAL] {report.get('report_number', '')} — {severity_label} Incident Reported",
            "html": html,
        }
        await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Critical incident notification sent to {len(recipient_emails)} recipients for {report.get('report_number')}")
    except Exception as e:
        logger.error(f"Failed to send critical incident notification: {e}")


async def _send_escalation_email(report: dict, admin_emails: list):
    """Send escalation email for reports not reviewed within 24 hours."""
    try:
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px;">
            <div style="text-align:center;padding:16px 0;">
                <h1 style="color:#7c3aed;margin:0;">Munal AI</h1>
            </div>
            <div style="background:#fffbeb;border-left:4px solid #f59e0b;border-radius:8px;padding:24px;margin:16px 0;">
                <h2 style="color:#92400e;margin-top:0;">Escalation — Unreviewed Report</h2>
                <p style="color:#4b5563;">The following report has been <strong>open for more than 24 hours</strong> without review:</p>
                <table style="width:100%;border-collapse:collapse;margin:16px 0;">
                    <tr><td style="padding:6px 0;color:#6b7280;width:130px;">Report #</td><td style="padding:6px 0;font-weight:bold;">{report.get('report_number', 'N/A')}</td></tr>
                    <tr><td style="padding:6px 0;color:#6b7280;">Severity</td><td style="padding:6px 0;">{report.get('severity', '').replace('_', ' ').title()}</td></tr>
                    <tr><td style="padding:6px 0;color:#6b7280;">Location</td><td style="padding:6px 0;">{report.get('location', '')}</td></tr>
                    <tr><td style="padding:6px 0;color:#6b7280;">Created</td><td style="padding:6px 0;">{report.get('created_at', '')}</td></tr>
                </table>
                <p style="color:#4b5563;">Please assign an investigator and update the report status.</p>
            </div>
            <div style="text-align:center;padding:16px 0;border-top:1px solid #e5e7eb;">
                <p style="color:#9ca3af;font-size:11px;">&copy; 2026 Munal AI. All rights reserved.</p>
            </div>
        </div>
        """

        params = {
            "from": SENDER_EMAIL,
            "to": admin_emails,
            "subject": f"[ESCALATION] Report {report.get('report_number', '')} — Unreviewed for 24+ hours",
            "html": html,
        }
        await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Escalation email sent for report {report.get('report_number')}")
    except Exception as e:
        logger.error(f"Failed to send escalation email: {e}")


# ============== Escalation Scheduler ==============

async def check_escalations():
    """Check for reports that have been 'open' for more than 24 hours and escalate."""
    try:
        cutoff = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
        stale_reports = await db.incident_reports.find(
            {
                "status": "open",
                "created_at": {"$lte": cutoff},
                "escalated": {"$ne": True},
            },
            {"_id": 0}
        ).to_list(100)

        if not stale_reports:
            return

        admins = await db.users.find(
            {"role": "Admin"},
            {"_id": 0, "email": 1}
        ).to_list(50)
        admin_emails = [a["email"] for a in admins if a.get("email")]

        if not admin_emails:
            logger.info("No admin emails found for escalation")
            return

        for report in stale_reports:
            await _send_escalation_email(report, admin_emails)
            # Mark as escalated so we don't re-send
            await db.incident_reports.update_one(
                {"id": report["id"]},
                {
                    "$set": {"escalated": True, "escalated_at": datetime.now(timezone.utc).isoformat()},
                    "$push": {"audit_log": {
                        "action": "escalated",
                        "by": "system",
                        "by_name": "Auto-Escalation",
                        "at": datetime.now(timezone.utc).isoformat(),
                    }}
                }
            )

        logger.info(f"Escalation check complete: {len(stale_reports)} reports escalated")
    except Exception as e:
        logger.error(f"Escalation check failed: {e}")
