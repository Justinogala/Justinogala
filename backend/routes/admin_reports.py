"""
Admin Report Generation — Real-time reports from MongoDB.
Generates PDF/Excel for: User Activity, Meeting Summary, System Performance,
Security Audit, Storage Usage, Revenue & Billing, Subscriptions.
"""
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone, timedelta
from typing import Optional
import io

from config import db, logger

router = APIRouter(prefix="/admin/reports", tags=["admin-reports"])


def _safe_str(val, max_len=None):
    """Convert any value to string safely."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.isoformat()
    s = str(val)
    if max_len:
        return s[:max_len]
    return s


def _safe_date(val):
    """Extract date portion from any value."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val)[:10]


def _date_filter(start_date: str, end_date: str, field: str = "timestamp"):
    q = {}
    if start_date:
        q.setdefault(field, {})["$gte"] = start_date
    if end_date:
        q.setdefault(field, {})["$lte"] = end_date + "T23:59:59"
    return q


# ============ Data Fetchers ============

async def _user_activity_data(start_date, end_date):
    date_q = _date_filter(start_date, end_date, "timestamp")
    activities = await db.user_activity.find({**date_q}, {"_id": 0}).sort("timestamp", -1).to_list(2000)
    sessions = await db.user_sessions.find({**_date_filter(start_date, end_date, "started_at")}, {"_id": 0}).to_list(2000)
    users = await db.users.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1, "role": 1, "created_at": 1, "status": 1}).to_list(500)

    # Aggregate by action
    action_counts = {}
    for a in activities:
        act = a.get("action", "unknown")
        action_counts[act] = action_counts.get(act, 0) + 1

    # Daily activity
    daily = {}
    for a in activities:
        day = _safe_date(a.get("timestamp"))
        if day:
            daily[day] = daily.get(day, 0) + 1

    # New users in period
    new_users = [u for u in users if start_date and _safe_str(u.get("created_at")) >= start_date]

    return {
        "title": "User Activity Report",
        "summary": {
            "Total Activities": len(activities),
            "Unique Sessions": len(sessions),
            "Active Sessions": sum(1 for s in sessions if s.get("active")),
            "New Users (Period)": len(new_users),
            "Total Users": len(users),
        },
        "table_headers": ["Timestamp", "User ID", "Action", "Details", "IP Address"],
        "table_rows": [
            [_safe_str(a.get("timestamp")), a.get("user_id", ""), a.get("action", ""), _safe_str(a.get("details"), 80), a.get("ip_address", "")]
            for a in activities[:500]
        ],
        "charts": {"action_breakdown": action_counts, "daily_activity": dict(sorted(daily.items()))},
    }


async def _meeting_summary_data(start_date, end_date):
    date_q = _date_filter(start_date, end_date, "start_time")
    events = await db.calendar_events.find({**date_q}, {"_id": 0}).sort("start_time", -1).to_list(2000)

    # Breakdown by category
    cat_counts = {}
    for e in events:
        cat = e.get("category", "uncategorized")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    # Recurrence stats
    recurring = sum(1 for e in events if e.get("recurrence") and e["recurrence"] != "none")

    return {
        "title": "Meeting Summary Report",
        "summary": {
            "Total Events": len(events),
            "Recurring Events": recurring,
            "One-time Events": len(events) - recurring,
            "Categories": len(cat_counts),
        },
        "table_headers": ["Date", "Title", "Category", "Location", "Start", "End", "Recurrence"],
        "table_rows": [
            [_safe_date(e.get("start_time")), e.get("title", ""), e.get("category", ""), e.get("location", ""), _safe_str(e.get("start_time"))[11:16] if len(_safe_str(e.get("start_time"))) > 11 else "", _safe_str(e.get("end_time"))[11:16] if len(_safe_str(e.get("end_time"))) > 11 else "", e.get("recurrence", "none")]
            for e in events[:500]
        ],
        "charts": {"category_breakdown": cat_counts},
    }


async def _system_performance_data(start_date, end_date):
    sessions = await db.user_sessions.find({**_date_filter(start_date, end_date, "started_at")}, {"_id": 0}).to_list(2000)
    users_count = await db.users.count_documents({})
    collections = await db.list_collection_names()

    # Daily session counts
    daily_sessions = {}
    for s in sessions:
        day = _safe_date(s.get("started_at"))
        if day:
            daily_sessions[day] = daily_sessions.get(day, 0) + 1

    col_stats = []
    for c in sorted(collections):
        if c.endswith(".chunks"):
            continue
        count = await db[c].count_documents({})
        col_stats.append({"collection": c, "documents": count})

    return {
        "title": "System Performance Report",
        "summary": {
            "Total Collections": len([c for c in collections if not c.endswith(".chunks")]),
            "Total Users": users_count,
            "Sessions (Period)": len(sessions),
            "Active Sessions": sum(1 for s in sessions if s.get("active")),
            "Peak Daily Sessions": max(daily_sessions.values()) if daily_sessions else 0,
        },
        "table_headers": ["Collection", "Document Count"],
        "table_rows": [[c["collection"], str(c["documents"])] for c in col_stats],
        "charts": {"daily_sessions": dict(sorted(daily_sessions.items()))},
    }


async def _security_audit_data(start_date, end_date):
    date_q = _date_filter(start_date, end_date, "timestamp")
    logs = await db.audit_logs.find({**date_q}, {"_id": 0}).sort("timestamp", -1).to_list(2000)
    sessions = await db.user_sessions.find({**_date_filter(start_date, end_date, "started_at")}, {"_id": 0}).to_list(2000)

    # Category breakdown
    cat_counts = {}
    for entry in logs:
        cat = entry.get("category", "general")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    # Unique IPs
    ips = set()
    for entry in logs:
        ip = entry.get("ip_address")
        if ip:
            ips.add(ip)
    for s in sessions:
        ip = s.get("ip_address")
        if ip:
            ips.add(ip)

    return {
        "title": "Security Audit Report",
        "summary": {
            "Total Audit Events": len(logs),
            "Sessions (Period)": len(sessions),
            "Unique IP Addresses": len(ips),
            "Event Categories": len(cat_counts),
        },
        "table_headers": ["Timestamp", "Action", "Category", "Admin Email", "IP Address", "Details"],
        "table_rows": [
            [_safe_str(entry.get("timestamp")), entry.get("action", ""), entry.get("category", ""), entry.get("admin_email", ""), entry.get("ip_address", ""), _safe_str(entry.get("details"), 80)]
            for entry in logs[:500]
        ],
        "charts": {"category_breakdown": cat_counts},
    }


async def _storage_usage_data(start_date, end_date):
    chat_files = await db["chat_files.files"].find({}, {"_id": 0, "filename": 1, "length": 1, "uploadDate": 1, "metadata": 1}).to_list(500)
    msg_files = await db["message_attachments.files"].find({}, {"_id": 0, "filename": 1, "length": 1, "uploadDate": 1, "metadata": 1}).to_list(500)

    all_files = []
    total_size = 0
    for f in chat_files:
        size = f.get("length", 0)
        total_size += size
        all_files.append({"name": f.get("filename", ""), "size": size, "type": "Chat File", "date": _safe_date(f.get("uploadDate"))})
    for f in msg_files:
        size = f.get("length", 0)
        total_size += size
        all_files.append({"name": f.get("filename", ""), "size": size, "type": "Message Attachment", "date": _safe_date(f.get("uploadDate"))})

    # By type breakdown
    type_sizes = {}
    for f in all_files:
        t = f["type"]
        type_sizes[t] = type_sizes.get(t, 0) + f["size"]

    return {
        "title": "Storage Usage Report",
        "summary": {
            "Total Files": len(all_files),
            "Chat Files": len(chat_files),
            "Message Attachments": len(msg_files),
            "Total Size": f"{total_size / (1024*1024):.2f} MB",
        },
        "table_headers": ["Filename", "Type", "Size (KB)", "Upload Date"],
        "table_rows": [
            [f["name"], f["type"], f"{f['size']/1024:.1f}", f["date"]]
            for f in all_files
        ],
        "charts": {"storage_by_type": {k: round(v / 1024, 1) for k, v in type_sizes.items()}},
    }


async def _revenue_billing_data(start_date, end_date):
    date_q = _date_filter(start_date, end_date, "created_at")
    txns = await db.payment_transactions.find({**date_q}, {"_id": 0}).sort("created_at", -1).to_list(5000)

    total_revenue = sum(t.get("amount", 0) for t in txns if t.get("payment_status") == "completed")
    completed = [t for t in txns if t.get("payment_status") == "completed"]
    pending = [t for t in txns if t.get("payment_status") == "pending"]

    # Revenue by package
    by_package = {}
    for t in completed:
        pkg = t.get("package_name", "Unknown")
        by_package[pkg] = by_package.get(pkg, 0) + t.get("amount", 0)

    # Daily revenue
    daily_rev = {}
    for t in completed:
        day = _safe_date(t.get("created_at"))
        if day:
            daily_rev[day] = daily_rev.get(day, 0) + t.get("amount", 0)

    return {
        "title": "Revenue & Billing Report",
        "summary": {
            "Total Transactions": len(txns),
            "Completed": len(completed),
            "Pending": len(pending),
            "Total Revenue": f"${total_revenue/100:.2f}" if total_revenue > 100 else f"${total_revenue:.2f}",
        },
        "table_headers": ["Date", "User Email", "Package", "Amount", "Currency", "Status"],
        "table_rows": [
            [_safe_date(t.get("created_at")), t.get("user_email", ""), t.get("package_name", ""), str(t.get("amount", 0)), (t.get("currency") or "usd").upper(), t.get("payment_status", "")]
            for t in txns[:500]
        ],
        "charts": {"revenue_by_package": by_package, "daily_revenue": dict(sorted(daily_rev.items()))},
    }


async def _subscriptions_data(start_date, end_date):
    plans = await db.subscription_plans.find({}, {"_id": 0}).to_list(50)
    users = await db.users.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1, "plan": 1, "status": 1, "created_at": 1}).to_list(5000)

    # Users by plan
    by_plan = {}
    for u in users:
        p = u.get("plan", "Free")
        by_plan[p] = by_plan.get(p, 0) + 1

    # Active vs inactive
    active = sum(1 for u in users if u.get("status") in ("Active", "active", None))

    return {
        "title": "Subscriptions Report",
        "summary": {
            "Total Plans": len(plans),
            "Total Users": len(users),
            "Active Users": active,
            "Inactive Users": len(users) - active,
        },
        "table_headers": ["Plan Name", "Price (Monthly)", "Price (Annual)", "Active", "Subscribers"],
        "table_rows": [
            [p.get("name", ""), str(p.get("price_monthly", 0)), str(p.get("price_annual", 0)), "Yes" if p.get("is_active") else "No", str(p.get("subscribers", 0))]
            for p in plans
        ],
        "extra_table": {
            "title": "Users by Plan",
            "headers": ["Email", "Name", "Plan", "Status", "Joined"],
            "rows": [
                [u.get("email", ""), u.get("name", ""), u.get("plan", "Free"), u.get("status", "Active"), _safe_date(u.get("created_at"))]
                for u in users[:200]
            ],
        },
        "charts": {"users_by_plan": by_plan},
    }


REPORT_FETCHERS = {
    "User Activity": _user_activity_data,
    "Meeting Summary": _meeting_summary_data,
    "System Performance": _system_performance_data,
    "Security Audit": _security_audit_data,
    "Storage Usage": _storage_usage_data,
    "Revenue & Billing": _revenue_billing_data,
    "Subscriptions": _subscriptions_data,
}


# ============ PDF Generator ============

def _build_pdf(data: dict, start_date: str, end_date: str):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    story = []

    title_s = ParagraphStyle('T', parent=styles['Title'], fontSize=20, textColor=colors.HexColor('#1e1b4b'), spaceAfter=4)
    sub_s = ParagraphStyle('S', parent=styles['Normal'], fontSize=10, textColor=colors.grey, spaceAfter=14)
    section_s = ParagraphStyle('Sec', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#4338ca'), spaceBefore=14, spaceAfter=6)

    # Header
    story.append(Paragraph(data["title"], title_s))
    story.append(Paragraph(f"{start_date or 'All time'} to {end_date or 'present'} &bull; Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", sub_s))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e5e7eb')))
    story.append(Spacer(1, 6))

    # Summary cards as table
    summary = data.get("summary", {})
    if summary:
        story.append(Paragraph("Summary", section_s))
        sum_rows = [list(summary.keys()), [str(v) for v in summary.values()]]
        col_w = min(120, 460 / max(1, len(summary)))
        t = Table(sum_rows, colWidths=[col_w] * len(summary))
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4338ca')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f3f4f6')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 11),
        ]))
        story.append(t)

    # Main data table
    headers = data.get("table_headers", [])
    rows = data.get("table_rows", [])
    if headers and rows:
        story.append(Paragraph(f"Detail Records ({len(rows)} rows)", section_s))
        table_data = [headers] + rows[:200]  # Cap at 200 for PDF
        n_cols = len(headers)
        col_w = min(100, 480 / max(1, n_cols))
        t = Table(table_data, colWidths=[col_w] * n_cols, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4338ca')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e5e7eb')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(t)

    # Extra table (e.g. Subscriptions has users_by_plan)
    extra = data.get("extra_table")
    if extra:
        story.append(Paragraph(extra["title"], section_s))
        et_data = [extra["headers"]] + extra["rows"][:100]
        n_cols = len(extra["headers"])
        col_w = min(100, 480 / max(1, n_cols))
        t = Table(et_data, colWidths=[col_w] * n_cols, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e5e7eb')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(t)

    # Footer
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#d1d5db')))
    story.append(Paragraph("Munal AI — Confidential", ParagraphStyle('F', parent=styles['Normal'], fontSize=7, textColor=colors.lightgrey, spaceBefore=4)))

    doc.build(story)
    buf.seek(0)
    return buf


# ============ Excel Generator ============

def _build_excel(data: dict, start_date: str, end_date: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    ws_sum.cell(row=1, column=1, value=data["title"]).font = Font(bold=True, size=14)
    ws_sum.cell(row=2, column=1, value=f"{start_date or 'All time'} to {end_date or 'present'}")
    ws_sum.cell(row=3, column=1, value=f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")

    row = 5
    summary = data.get("summary", {})
    for key, val in summary.items():
        ws_sum.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws_sum.cell(row=row, column=2, value=str(val))
        row += 1

    # Charts data
    charts = data.get("charts", {})
    if charts:
        row += 1
        for chart_name, chart_data in charts.items():
            ws_sum.cell(row=row, column=1, value=chart_name.replace("_", " ").title()).font = Font(bold=True, size=11)
            row += 1
            if isinstance(chart_data, dict):
                for k, v in chart_data.items():
                    ws_sum.cell(row=row, column=1, value=str(k))
                    ws_sum.cell(row=row, column=2, value=v)
                    row += 1
            row += 1

    # Auto-width summary
    for col in ws_sum.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 40))
        ws_sum.column_dimensions[col[0].column_letter].width = max_len + 3

    # Data sheet
    headers = data.get("table_headers", [])
    rows = data.get("table_rows", [])
    if headers and rows:
        ws_data = wb.create_sheet("Data")
        for col_num, h in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row_num, row_data in enumerate(rows, 2):
            for col_num, val in enumerate(row_data, 1):
                cell = ws_data.cell(row=row_num, column=col_num, value=str(val) if val else "")
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col in ws_data.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 35))
            ws_data.column_dimensions[col[0].column_letter].width = max_len + 3

    # Extra table sheet
    extra = data.get("extra_table")
    if extra:
        ws_extra = wb.create_sheet(extra["title"])
        for col_num, h in enumerate(extra["headers"], 1):
            cell = ws_extra.cell(row=1, column=col_num, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
            cell.border = thin_border
        for row_num, row_data in enumerate(extra["rows"], 2):
            for col_num, val in enumerate(row_data, 1):
                cell = ws_extra.cell(row=row_num, column=col_num, value=str(val) if val else "")
                cell.border = thin_border

        for col in ws_extra.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 35))
            ws_extra.column_dimensions[col[0].column_letter].width = max_len + 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============ Endpoint ============

@router.get("/generate")
async def generate_admin_report(
    type: str = Query(..., description="Report type"),
    format: str = Query("PDF", description="PDF or Excel"),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    """Generate a real-time admin report from MongoDB data."""
    fetcher = REPORT_FETCHERS.get(type)
    if not fetcher:
        raise HTTPException(status_code=400, detail=f"Unknown report type: {type}. Available: {list(REPORT_FETCHERS.keys())}")

    try:
        data = await fetcher(start_date or "", end_date or "")

        if format.lower() == "excel":
            buf = _build_excel(data, start_date, end_date)
            safe_name = type.lower().replace(" ", "_").replace("&", "and")
            filename = f"{safe_name}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )
        else:
            buf = _build_pdf(data, start_date, end_date)
            safe_name = type.lower().replace(" ", "_").replace("&", "and")
            filename = f"{safe_name}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.pdf"
            return StreamingResponse(
                buf,
                media_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )
    except Exception as e:
        logger.error(f"Error generating admin report '{type}': {e}")
        raise HTTPException(status_code=500, detail=str(e))
