"""
Sheets API — AI-powered spreadsheet intelligence for Munal Workplace.
Provides CRUD for spreadsheets and AI generation capabilities.
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field
from typing import Optional, List, Any
from datetime import datetime, timezone
from config import db
import uuid
import json
import os
import io

router = APIRouter(prefix="/sheets", tags=["sheets"])

# ── Auth dependency (reuse from auth module) ──
from routes.auth import get_current_user

EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY", "")


# ── Models ──
class CreateSheetRequest(BaseModel):
    title: str = "Untitled Spreadsheet"
    workspace_id: Optional[str] = None

class UpdateSheetRequest(BaseModel):
    title: Optional[str] = None
    data: Optional[List[Any]] = None  # Fortune-Sheet data array
    
class AIGenerateRequest(BaseModel):
    prompt: str
    sheet_id: Optional[str] = None

class AIFormulaRequest(BaseModel):
    description: str
    context: Optional[str] = None  # surrounding cell values for context


# ── CRUD Endpoints ──

@router.post("")
async def create_sheet(req: CreateSheetRequest, user: dict = Depends(get_current_user)):
    sheet = {
        "id": str(uuid.uuid4()),
        "title": req.title,
        "workspace_id": req.workspace_id,
        "created_by": user["id"],
        "created_by_name": user.get("full_name", user.get("email", "")),
        "data": [{"name": "Sheet1", "celldata": [], "order": 0, "row": 50, "column": 26, "status": 1}],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.sheets.insert_one(sheet)
    sheet.pop("_id", None)
    return sheet


@router.get("")
async def list_sheets(
    workspace_id: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {"created_by": user["id"]}
    if workspace_id:
        query["workspace_id"] = workspace_id
    sheets = await db.sheets.find(query, {"_id": 0, "data": 0}).sort("updated_at", -1).to_list(200)
    return sheets


def _ensure_celldata(sheet_data: list) -> list:
    """Convert 2D data arrays to celldata (sparse) format for Fortune-Sheet compatibility."""
    for s in sheet_data:
        if "data" in s and isinstance(s["data"], list) and len(s["data"]) > 0:
            # Check if it's a 2D array (not celldata)
            if isinstance(s["data"][0], list):
                celldata = []
                for ri, row in enumerate(s["data"]):
                    if not isinstance(row, list):
                        continue
                    for ci, cell in enumerate(row):
                        if cell is not None:
                            celldata.append({"r": ri, "c": ci, "v": cell})
                s["celldata"] = celldata
                del s["data"]
    return sheet_data


@router.get("/{sheet_id}")
async def get_sheet(sheet_id: str, user: dict = Depends(get_current_user)):
    sheet = await db.sheets.find_one({"id": sheet_id, "created_by": user["id"]}, {"_id": 0})
    if not sheet:
        raise HTTPException(404, "Sheet not found")
    if "data" in sheet:
        sheet["data"] = _ensure_celldata(sheet["data"])
    return sheet


@router.put("/{sheet_id}")
async def update_sheet(sheet_id: str, req: UpdateSheetRequest, user: dict = Depends(get_current_user)):
    update = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if req.title is not None:
        update["title"] = req.title
    if req.data is not None:
        update["data"] = req.data
    result = await db.sheets.update_one(
        {"id": sheet_id, "created_by": user["id"]},
        {"$set": update}
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Sheet not found")
    return {"status": "ok"}


@router.delete("/{sheet_id}")
async def delete_sheet(sheet_id: str, user: dict = Depends(get_current_user)):
    result = await db.sheets.delete_one({"id": sheet_id, "created_by": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Sheet not found")
    return {"status": "deleted"}


@router.post("/{sheet_id}/duplicate")
async def duplicate_sheet(sheet_id: str, user: dict = Depends(get_current_user)):
    original = await db.sheets.find_one({"id": sheet_id, "created_by": user["id"]}, {"_id": 0})
    if not original:
        raise HTTPException(404, "Sheet not found")
    new_sheet = {
        **original,
        "id": str(uuid.uuid4()),
        "title": f"{original['title']} (Copy)",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.sheets.insert_one(new_sheet)
    new_sheet.pop("_id", None)
    return new_sheet


# ── AI Endpoints ──

@router.post("/ai/generate")
async def ai_generate_sheet(req: AIGenerateRequest, user: dict = Depends(get_current_user)):
    """Generate a spreadsheet from a natural language prompt."""
    if not EMERGENT_KEY:
        raise HTTPException(500, "AI service not configured")

    from llm_client import chat_completion

    system = """You are a spreadsheet generator for a business platform. When given a prompt, generate structured spreadsheet data.

Return ONLY valid JSON in this exact format (no markdown, no explanation):
{
  "title": "Sheet Title",
  "columns": ["Column A Header", "Column B Header", ...],
  "rows": [
    ["row1_val1", "row1_val2", ...],
    ["row2_val1", "row2_val2", ...]
  ],
  "column_widths": [120, 120, ...],
  "formulas": [
    {"r": 10, "c": 1, "f": "=SUM(B2:B10)"}
  ]
}

Rules:
- Generate 5-20 realistic sample rows with realistic data
- Include formulas where appropriate (SUM, AVERAGE, COUNT, etc.)
- Column widths should reflect content (narrow for dates/numbers, wider for text)
- Use business-appropriate formatting
- Numbers should be numbers, not strings
- Dates should be formatted as YYYY-MM-DD"""

    try:
        result = chat_completion(
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": req.prompt},
            ],
            model="gpt-5.2",
            api_key=EMERGENT_KEY,
        )
        raw = result.choices[0].message.content.strip()
        # Strip markdown code fences if present
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
        if raw.endswith("```"):
            raw = raw[:-3]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

        ai_data = json.loads(raw)

        # Convert AI output to Fortune-Sheet celldata format
        celldata = []
        columns = ai_data.get("columns", [])
        rows = ai_data.get("rows", [])
        col_widths = ai_data.get("column_widths", [120] * len(columns))

        # Header row (bold, with background)
        for ci, col in enumerate(columns):
            celldata.append({
                "r": 0, "c": ci,
                "v": {
                    "v": col, "m": str(col), "ct": {"fa": "General", "t": "g"},
                    "bl": 1,  # bold
                    "bg": "#f0f4ff",
                    "fc": "#1a1a2e",
                }
            })

        # Data rows
        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                cell = {"r": ri + 1, "c": ci, "v": {}}
                if isinstance(val, (int, float)):
                    cell["v"] = {"v": val, "m": str(val), "ct": {"fa": "General", "t": "n"}}
                else:
                    cell["v"] = {"v": str(val), "m": str(val), "ct": {"fa": "General", "t": "g"}}
                celldata.append(cell)

        # Formulas
        for f in ai_data.get("formulas", []):
            celldata.append({
                "r": f["r"], "c": f["c"],
                "v": {"f": f["f"], "v": 0, "m": "0", "ct": {"fa": "General", "t": "n"}, "bl": 1, "bg": "#fffde7"}
            })

        # Build column config
        col_config = {str(i): w for i, w in enumerate(col_widths)}

        sheet_data = [{
            "name": "Sheet1",
            "celldata": celldata,
            "order": 0,
            "row": max(len(rows) + 5, 50),
            "column": max(len(columns) + 2, 26),
            "status": 1,
            "config": {
                "columnlen": col_config,
                "rowhidden": {},
                "colhidden": {},
            }
        }]

        title = ai_data.get("title", "AI Generated Sheet")

        # Save if sheet_id provided, else create new
        if req.sheet_id:
            await db.sheets.update_one(
                {"id": req.sheet_id, "created_by": user["id"]},
                {"$set": {"data": sheet_data, "title": title, "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
            return {"sheet_id": req.sheet_id, "title": title, "data": sheet_data}
        else:
            sheet = {
                "id": str(uuid.uuid4()),
                "title": title,
                "workspace_id": None,
                "created_by": user["id"],
                "created_by_name": user.get("full_name", user.get("email", "")),
                "data": sheet_data,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.sheets.insert_one(sheet)
            sheet.pop("_id", None)
            return sheet

    except json.JSONDecodeError:
        raise HTTPException(422, "AI returned invalid data. Please try rephrasing your prompt.")
    except Exception as e:
        raise HTTPException(500, f"AI generation failed: {str(e)}")


@router.post("/ai/formula")
async def ai_formula(req: AIFormulaRequest, user: dict = Depends(get_current_user)):
    """Convert natural language to spreadsheet formula."""
    if not EMERGENT_KEY:
        raise HTTPException(500, "AI service not configured")

    from llm_client import chat_completion

    system = """You are a spreadsheet formula expert. Convert natural language descriptions to spreadsheet formulas.
Return ONLY the formula string starting with '='. No explanation, no markdown.
Use standard Excel/Sheets formula syntax (SUM, AVERAGE, IF, VLOOKUP, COUNTIF, etc.).
Column references use letters (A, B, C...), row references use numbers (1, 2, 3...)."""

    context_msg = ""
    if req.context:
        context_msg = f"\n\nContext of surrounding cells:\n{req.context}"

    result = chat_completion(
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": f"{req.description}{context_msg}"},
        ],
        model="gpt-5.2",
        api_key=EMERGENT_KEY,
    )
    formula = result.choices[0].message.content.strip()
    if not formula.startswith("="):
        formula = "=" + formula
    return {"formula": formula}


# ── Phase 2: Chat with Data ──

class ChatWithDataRequest(BaseModel):
    message: str
    sheet_data_summary: Optional[str] = None  # frontend sends a compact representation

@router.post("/{sheet_id}/ai/chat")
async def chat_with_data(sheet_id: str, req: ChatWithDataRequest, user: dict = Depends(get_current_user)):
    """Chat with AI about the active sheet data."""
    if not EMERGENT_KEY:
        raise HTTPException(500, "AI service not configured")

    from llm_client import chat_completion

    # Load sheet data for context
    sheet = await db.sheets.find_one({"id": sheet_id, "created_by": user["id"]}, {"_id": 0})
    if not sheet:
        raise HTTPException(404, "Sheet not found")

    # Build a compact data summary from celldata
    data_context = req.sheet_data_summary or _extract_data_summary(sheet.get("data", []))

    system = f"""You are Munal AI, a data analyst assistant embedded in a spreadsheet application.
You have access to the user's spreadsheet data below. Answer questions about it accurately and concisely.
When referencing cells, use standard notation (A1, B2, etc.). Provide insights, calculations, and suggestions.
If the user asks for a formula, provide it. If they ask for analysis, be specific with numbers.

SPREADSHEET DATA:
{data_context}"""

    result = chat_completion(
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": req.message},
        ],
        model="gpt-5.2",
        api_key=EMERGENT_KEY,
    )
    return {"response": result.choices[0].message.content}


def _extract_data_summary(data: list) -> str:
    """Extract a readable summary from Fortune-Sheet data (supports both celldata and 2D array formats)."""
    if not data:
        return "Empty spreadsheet"
    
    lines = []
    for sheet_obj in data[:1]:  # first sheet only
        # Try celldata format first (sparse format: [{r, c, v}, ...])
        celldata = sheet_obj.get("celldata", [])
        # Also try 2D array format (dense format: [[cell, cell, ...], ...])
        data_2d = sheet_obj.get("data", [])
        
        rows = {}
        
        if celldata:
            # Handle sparse celldata format
            for cell in celldata:
                r, c = cell.get("r", 0), cell.get("c", 0)
                v = cell.get("v", {})
                val = v.get("m") or v.get("v") or ""
                if val:
                    rows.setdefault(r, {})[c] = str(val)
        elif data_2d and isinstance(data_2d, list):
            # Handle 2D array format from Fortune-Sheet
            for r_idx, row in enumerate(data_2d[:30]):  # limit to 30 rows
                if not row or not isinstance(row, list):
                    continue
                for c_idx, cell in enumerate(row):
                    if cell is None:
                        continue
                    if isinstance(cell, dict):
                        val = cell.get("m") or cell.get("v") or ""
                    else:
                        val = str(cell) if cell else ""
                    if val:
                        rows.setdefault(r_idx, {})[c_idx] = str(val)
        
        if not rows:
            continue
            
        # Format as table
        max_col = max((max(cols.keys()) for cols in rows.values() if cols), default=0)
        for r_idx in sorted(rows.keys())[:30]:  # limit to 30 rows
            row_vals = [rows[r_idx].get(c, "") for c in range(max_col + 1)]
            prefix = "HEADER: " if r_idx == 0 else f"Row {r_idx}: "
            lines.append(prefix + " | ".join(row_vals))
    
    return "\n".join(lines) if lines else "Empty spreadsheet"


# ── Phase 2: Smart Automation ──

class AutofillRequest(BaseModel):
    column_index: int
    column_name: str
    existing_values: List[str]  # values already in the column
    row_count: int = 10  # how many to generate
    context_columns: Optional[dict] = None  # adjacent column data for context

class SmartActionRequest(BaseModel):
    action: str  # "summarize", "sentiment", "categorize", "translate"
    values: List[str]
    options: Optional[dict] = None

@router.post("/{sheet_id}/ai/autofill")
async def ai_autofill(sheet_id: str, req: AutofillRequest, user: dict = Depends(get_current_user)):
    """AI auto-fill a column based on patterns and context."""
    if not EMERGENT_KEY:
        raise HTTPException(500, "AI service not configured")

    from llm_client import chat_completion

    context = f"Column: {req.column_name}\nExisting values: {', '.join(req.existing_values[:20])}"
    if req.context_columns:
        context += "\nAdjacent columns:\n"
        for col_name, vals in list(req.context_columns.items())[:5]:
            context += f"  {col_name}: {', '.join(str(v) for v in vals[:10])}\n"

    system = f"""You are a spreadsheet auto-fill assistant. Based on the pattern of existing values and context, generate {req.row_count} new values that follow the same pattern.

Return ONLY a JSON array of strings. No explanation, no markdown.
Example: ["value1", "value2", "value3"]

{context}"""

    result = chat_completion(
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": f"Generate {req.row_count} more values for the '{req.column_name}' column following the pattern."},
        ],
        model="gpt-5.2",
        api_key=EMERGENT_KEY,
    )
    raw = result.choices[0].message.content.strip()
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
    if raw.endswith("```"):
        raw = raw[:-3]
    if raw.startswith("json"):
        raw = raw[4:]
    raw = raw.strip()
    
    try:
        values = json.loads(raw)
        if not isinstance(values, list):
            values = [str(values)]
    except json.JSONDecodeError:
        values = [raw]
    
    return {"values": [str(v) for v in values]}


@router.post("/{sheet_id}/ai/smart-action")
async def smart_action(sheet_id: str, req: SmartActionRequest, user: dict = Depends(get_current_user)):
    """Run a smart AI action on a list of cell values."""
    if not EMERGENT_KEY:
        raise HTTPException(500, "AI service not configured")

    from llm_client import chat_completion

    action_prompts = {
        "summarize": "Summarize each text value in 1-2 sentences. Return a JSON array of summary strings.",
        "sentiment": "Analyze the sentiment of each text. Return a JSON array of objects: [{\"text\": \"original\", \"sentiment\": \"positive/negative/neutral\", \"score\": 0.0-1.0}]",
        "categorize": "Categorize each text into a relevant category. Return a JSON array of category strings.",
        "translate": f"Translate each text to {req.options.get('target_language', 'English') if req.options else 'English'}. Return a JSON array of translated strings.",
    }

    prompt = action_prompts.get(req.action, f"Process each value with action '{req.action}'. Return a JSON array.")
    values_text = "\n".join(f"{i+1}. {v}" for i, v in enumerate(req.values[:50]))

    result = chat_completion(
        messages=[
            {"role": "system", "content": f"{prompt}\n\nReturn ONLY valid JSON array. No markdown, no explanation."},
            {"role": "user", "content": f"Process these values:\n{values_text}"},
        ],
        model="gpt-5.2",
        api_key=EMERGENT_KEY,
    )
    raw = result.choices[0].message.content.strip()
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
    if raw.endswith("```"):
        raw = raw[:-3]
    if raw.startswith("json"):
        raw = raw[4:]
    raw = raw.strip()

    try:
        results = json.loads(raw)
    except json.JSONDecodeError:
        results = [raw]

    return {"results": results, "action": req.action}


# ── Download Sheet as XLSX ──

@router.get("/{sheet_id}/download")
async def download_sheet(sheet_id: str, user: dict = Depends(get_current_user)):
    """Export a sheet as an .xlsx file."""
    from openpyxl import Workbook as XLWorkbook
    from openpyxl.styles import Font, PatternFill, Alignment

    sheet = await db.sheets.find_one({"id": sheet_id, "created_by": user["id"]}, {"_id": 0})
    if not sheet:
        raise HTTPException(404, "Sheet not found")

    wb = XLWorkbook()
    sheet_data_list = sheet.get("data", [])

    for idx, fs_sheet in enumerate(sheet_data_list):
        if idx == 0:
            ws = wb.active
            ws.title = fs_sheet.get("name", "Sheet1")
        else:
            ws = wb.create_sheet(title=fs_sheet.get("name", f"Sheet{idx+1}"))

        celldata = fs_sheet.get("celldata", [])
        config = fs_sheet.get("config", {})
        col_widths = config.get("columnlen", {})

        # Set column widths
        for col_str, width in col_widths.items():
            col_idx = int(col_str) + 1
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = max(width / 7, 8)

        for cell in celldata:
            r = cell.get("r", 0) + 1
            c = cell.get("c", 0) + 1
            v = cell.get("v", {})
            if not v:
                continue

            xc = ws.cell(row=r, column=c)

            # Handle formula
            if v.get("f"):
                xc.value = v["f"]
            elif v.get("ct", {}).get("t") == "n":
                try:
                    xc.value = float(v.get("v", 0))
                except (TypeError, ValueError):
                    xc.value = v.get("m", "")
            else:
                xc.value = v.get("m") or v.get("v") or ""

            # Bold
            is_bold = v.get("bl") == 1
            font_color = v.get("fc", "000000").lstrip("#") if v.get("fc") else "000000"
            if len(font_color) != 6:
                font_color = "000000"
            xc.font = Font(bold=is_bold, color=font_color)

            # Background
            if v.get("bg"):
                bg_color = v["bg"].lstrip("#")
                if len(bg_color) == 6:
                    xc.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = "".join(c for c in sheet.get("title", "Sheet") if c.isalnum() or c in " _-").strip() or "Sheet"
    filename = f"{safe_title}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Phase 3: AI Insights ──

class AIInsightsRequest(BaseModel):
    sheet_data_summary: Optional[str] = None

@router.post("/{sheet_id}/ai/insights")
async def ai_insights(sheet_id: str, req: AIInsightsRequest, user: dict = Depends(get_current_user)):
    """Analyze sheet data and return AI-generated insights."""
    if not EMERGENT_KEY:
        raise HTTPException(500, "AI service not configured")

    from llm_client import chat_completion

    sheet = await db.sheets.find_one({"id": sheet_id, "created_by": user["id"]}, {"_id": 0})
    if not sheet:
        raise HTTPException(404, "Sheet not found")

    data_context = req.sheet_data_summary or _extract_data_summary(sheet.get("data", []))

    system = f"""You are a data analyst AI embedded in a spreadsheet application.
Analyze the spreadsheet data below and produce insights.

Return ONLY valid JSON in this exact format (no markdown, no explanation):
{{
  "summary": "Brief 1-2 sentence overview of the dataset",
  "key_metrics": [
    {{"label": "Metric Name", "value": "123", "trend": "up|down|stable"}},
  ],
  "insights": [
    {{"title": "Insight title", "description": "Detailed explanation", "type": "trend|outlier|pattern|recommendation"}},
  ],
  "charts": [
    {{
      "title": "Chart Title",
      "type": "bar|line|pie",
      "xKey": "column_name_for_x_axis",
      "yKeys": ["column_name_for_y1"],
      "data": [
        {{"name": "Label1", "value1": 100}},
        {{"name": "Label2", "value1": 200}}
      ]
    }}
  ]
}}

Rules:
- Provide 2-5 key metrics with trend direction
- Provide 3-6 actionable insights
- Suggest 1-3 charts that best visualize the data
- Chart data should be derived directly from the spreadsheet data
- For pie charts, use "name" and "value" keys in data
- For bar/line charts, use "name" as x-axis label
- Keep chart data to max 15 data points for readability
- If data is insufficient for charts, return empty charts array

SPREADSHEET DATA:
{data_context}"""

    try:
        result = chat_completion(
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": "Analyze this spreadsheet data and provide insights with chart suggestions."},
            ],
            model="gpt-5.2",
            api_key=EMERGENT_KEY,
        )
        raw = result.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
        if raw.endswith("```"):
            raw = raw[:-3]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

        insights = json.loads(raw)
        return insights
    except json.JSONDecodeError:
        raise HTTPException(422, "AI returned invalid insights. Try again.")
    except Exception as e:
        raise HTTPException(500, f"AI insights failed: {str(e)}")
